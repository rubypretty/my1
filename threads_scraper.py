import argparse
import csv
import json
import os
import re
import time
from datetime import datetime, timedelta
from pathlib import Path

from common import build_driver
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


ENV_FILE = Path(".env")
DEFAULT_URL_TABLE = "url_table.xlsx"
DEFAULT_OUTPUT = "threads_posts.xlsx"
OUTPUT_TIME_FORMAT = "%Y.%m.%d %H:%M"


def load_local_env(path: Path = ENV_FILE) -> None:
    if not path.exists():
        return

    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue

        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def resolve_path(path_text: str) -> Path:
    path = Path(path_text)
    if path.is_absolute():
        return path
    return Path.cwd() / path


def read_urls_from_excel(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
    url_col = header.index("url") if "url" in header else 0

    urls = []
    for row in rows[1:]:
        if not row or len(row) <= url_col or row[url_col] is None:
            continue
        url = str(row[url_col]).strip()
        if url.startswith("http"):
            urls.append(url)
    return urls


def normalize_threads_url(url: str) -> str:
    if "/post/" in url:
        return url.replace("https://www.threads.com/", "https://www.threads.net/")
    return url


def wait_for_visible_content(driver: webdriver.Chrome, timeout: int = 30) -> None:
    wait = WebDriverWait(driver, timeout)
    wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
    wait.until(lambda current: len((current.find_element(By.TAG_NAME, "body").text or "").strip()) > 80)


def detect_unavailable_public_page(driver: webdriver.Chrome) -> str | None:
    body_text = driver.find_element(By.TAG_NAME, "body").text or ""
    current_url = driver.current_url

    if "instagram.com/accounts/login" in current_url:
        return "Redirected to Instagram login."
    if "此個人檔案不公開" in body_text or "This profile is private" in body_text:
        return "This Threads profile is private."
    if "登入以查看更多" in body_text or "Log in to see more" in body_text:
        return "Threads is asking for login before showing more content."
    return None


def clean_lines(text: str) -> list[str]:
    stop_markers = [
        "登入即可查看更多回覆。",
        "登入或註冊 Threads",
        "© 2026",
        "Threads 使用條款",
    ]
    for marker in stop_markers:
        if marker in text:
            text = text.split(marker, 1)[0].strip()

    noise_lines = {"首頁", "串文", "翻譯", "·", "作者", "登入"}
    lines = []
    for line in text.splitlines():
        line = line.strip()
        if not line or line in noise_lines:
            continue
        lines.append(line)
    return lines


def extract_author_from_url(url: str) -> str:
    match = re.search(r"/@([^/?#]+)/post/", url)
    return match.group(1) if match else ""


def looks_like_metric(line: str) -> bool:
    return bool(re.fullmatch(r"\d[\d,]*", line))


def looks_like_time(line: str) -> bool:
    if re.fullmatch(r"\d+\s*(秒|分鐘|小時|天|週|個月|年)", line):
        return True
    if re.fullmatch(r"\d{4}-\d{1,2}-\d{1,2}", line):
        return True
    return line in {"昨天", "前天"}


def format_post_time(post_time: str, reference_time: datetime | None = None) -> str:
    if not post_time:
        return ""

    reference_time = reference_time or datetime.now().astimezone()

    absolute_match = re.fullmatch(r"(\d{4})-(\d{1,2})-(\d{1,2})", post_time)
    if absolute_match:
        year, month, day = (int(part) for part in absolute_match.groups())
        return datetime(year, month, day, tzinfo=reference_time.tzinfo).strftime(OUTPUT_TIME_FORMAT)

    relative_match = re.fullmatch(r"(\d+)\s*(秒|分鐘|小時|天|週|個月|年)", post_time)
    if relative_match:
        amount = int(relative_match.group(1))
        unit = relative_match.group(2)
        deltas = {
            "秒": timedelta(seconds=amount),
            "分鐘": timedelta(minutes=amount),
            "小時": timedelta(hours=amount),
            "天": timedelta(days=amount),
            "週": timedelta(weeks=amount),
            "個月": timedelta(days=amount * 30),
            "年": timedelta(days=amount * 365),
        }
        return (reference_time - deltas[unit]).strftime(OUTPUT_TIME_FORMAT)

    if post_time == "昨天":
        return (reference_time - timedelta(days=1)).strftime(OUTPUT_TIME_FORMAT)
    if post_time == "前天":
        return (reference_time - timedelta(days=2)).strftime(OUTPUT_TIME_FORMAT)

    return post_time


def drop_carousel_markers(lines: list[str]) -> list[str]:
    cleaned = []
    index = 0
    while index < len(lines):
        if (
            index + 2 < len(lines)
            and lines[index].isdigit()
            and lines[index + 1] == "/"
            and lines[index + 2].isdigit()
        ):
            index += 3
            continue
        cleaned.append(lines[index])
        index += 1
    return cleaned


def find_reply_start(lines: list[str], start: int) -> int | None:
    for index in range(start, len(lines)):
        line = lines[index]
        if line.startswith("回覆"):
            return index + 1
        if line in {"熱門", "查看動態"}:
            return index
    return None


def find_metric_start(lines: list[str], start: int, stop: int) -> int | None:
    for index in range(start, stop):
        if looks_like_metric(lines[index]):
            metric_count = 0
            for line in lines[index : min(stop, index + 5)]:
                if looks_like_metric(line):
                    metric_count += 1
            if metric_count >= 2:
                return index
    return None


def find_metric_end(lines: list[str], start: int, stop: int) -> int:
    index = start
    while index < stop and looks_like_metric(lines[index]):
        index += 1
    return index


def trim_ui_lines(lines: list[str]) -> list[str]:
    return [
        line
        for line in lines
        if line not in {"熱門", "查看動態", "相關串文"} and not line.startswith("回覆")
    ]


def parse_thread_lines(
    lines: list[str],
    source_url: str = "",
    reference_time: datetime | None = None,
) -> dict:
    if not lines:
        return {}

    lines = drop_carousel_markers(lines)
    author_from_url = extract_author_from_url(source_url)

    view_index = next(
        (index for index, line in enumerate(lines[:8]) if "瀏覽" in line or "views" in line.lower()),
        None,
    )
    view_count = lines[view_index] if view_index is not None else ""

    author_index = None
    if author_from_url:
        author_index = next((index for index, line in enumerate(lines) if line == author_from_url), None)
    if author_index is None:
        start = (view_index + 1) if view_index is not None else 0
        author_index = start if start < len(lines) else None

    author = author_from_url or (lines[author_index] if author_index is not None else "")

    post_time = ""
    time_index = None
    if author_index is not None:
        for index in range(author_index + 1, min(len(lines), author_index + 6)):
            if looks_like_time(lines[index]):
                post_time = lines[index]
                time_index = index
                break

    text_start = (time_index + 1) if time_index is not None else ((author_index + 1) if author_index is not None else 0)
    reply_start = find_reply_start(lines, text_start)
    main_stop = reply_start - 1 if reply_start is not None else len(lines)
    metric_start = find_metric_start(lines, text_start, main_stop)
    if metric_start is not None:
        main_stop = metric_start

    main_lines = trim_ui_lines(lines[text_start:main_stop])

    engagement_stop = reply_start - 1 if reply_start is not None else len(lines)
    metric_source_start = metric_start if metric_start is not None else text_start
    metric_end = find_metric_end(lines, metric_source_start, engagement_stop)
    metrics = [
        line
        for line in lines[metric_source_start:metric_end]
        if looks_like_metric(line)
    ]

    if reply_start is not None:
        extra_lines = trim_ui_lines(lines[reply_start:])
    elif metric_start is not None:
        extra_lines = trim_ui_lines(lines[metric_end:])
    else:
        extra_lines = []

    return {
        "author": author,
        "post_time": format_post_time(post_time, reference_time),
        "view_count": view_count,
        "main_text": "\n".join(main_lines).strip(),
        "engagement": " / ".join(metrics[:3]),
        "related_or_replies": "\n".join(extra_lines).strip(),
    }


def extract_visible_thread(driver: webdriver.Chrome, source_url: str) -> dict | None:
    scraped_at = datetime.now().astimezone()
    text = (driver.find_element(By.TAG_NAME, "body").text or "").strip()
    if not text:
        return None

    lines = clean_lines(text)
    cleaned_text = "\n".join(lines).strip()
    if len(cleaned_text) < 20:
        return None

    parsed = parse_thread_lines(lines, source_url=source_url, reference_time=scraped_at)
    return {
        "source_url": source_url,
        "final_url": driver.current_url,
        **parsed,
        "text": cleaned_text,
        "scraped_at": scraped_at.strftime("%Y-%m-%dT%H:%M:%S%z"),
    }


def scrape_one_url(driver: webdriver.Chrome, url: str, delay: float) -> dict:
    open_url = normalize_threads_url(url)
    print(f"Scraping: {url}")
    driver.get(open_url)
    time.sleep(delay)

    try:
        wait_for_visible_content(driver)
    except TimeoutException:
        return {
            "source_url": url,
            "final_url": driver.current_url,
            "error": "Timed out waiting for public content.",
            "scraped_at": time.strftime("%Y-%m-%dT%H:%M:%S%z"),
        }

    unavailable_reason = detect_unavailable_public_page(driver)
    if unavailable_reason:
        return {
            "source_url": url,
            "final_url": driver.current_url,
            "error": unavailable_reason,
            "scraped_at": time.strftime("%Y-%m-%dT%H:%M:%S%z"),
        }

    post = extract_visible_thread(driver, source_url=url)
    if post:
        return post

    return {
        "source_url": url,
        "final_url": driver.current_url,
        "error": "No readable public post text found.",
        "scraped_at": time.strftime("%Y-%m-%dT%H:%M:%S%z"),
    }


def scrape_urls(urls: list[str], delay: float, headless: bool) -> list[dict]:
    driver = build_driver(headless=headless)
    try:
        return [scrape_one_url(driver, url, delay=delay) for url in urls]
    finally:
        driver.quit()


def save_json(posts: list[dict], output: Path) -> None:
    output.write_text(json.dumps(posts, ensure_ascii=False, indent=2), encoding="utf-8")


def save_csv(posts: list[dict], output: Path) -> None:
    fieldnames = [
        "source_url",
        "final_url",
        "author",
        "post_time",
        "view_count",
        "main_text",
        "engagement",
        "related_or_replies",
        "error",
        "scraped_at",
    ]
    with output.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(posts)


def save_excel(posts: list[dict], output: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "posts"
    raw = workbook.create_sheet("raw")

    summary_headers = [
        "source_url",
        "author",
        "post_time",
        "view_count",
        "main_text",
        "engagement",
        "related_or_replies",
        "error",
        "scraped_at",
    ]
    raw_headers = ["source_url", "final_url", "text", "error", "scraped_at"]

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for sheet, headers in [(summary, summary_headers), (raw, raw_headers)]:
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for post in posts:
        summary.append([post.get(header, "") for header in summary_headers])
        raw.append([post.get(header, "") for header in raw_headers])

    for sheet in [summary, raw]:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column, width in {
        "A": 56,
        "B": 20,
        "C": 14,
        "D": 14,
        "E": 60,
        "F": 16,
        "G": 70,
        "H": 32,
        "I": 22,
    }.items():
        summary.column_dimensions[column].width = width

    raw.column_dimensions["A"].width = 56
    raw.column_dimensions["B"].width = 56
    raw.column_dimensions["C"].width = 100
    raw.column_dimensions["D"].width = 32
    raw.column_dimensions["E"].width = 22
    workbook.save(output)


def main() -> None:
    load_local_env()
    default_url_table = os.getenv("THREADS_URL_TABLE", DEFAULT_URL_TABLE)

    parser = argparse.ArgumentParser(description="Scrape public Threads post URLs from an Excel file")
    parser.add_argument("--url-table", default=default_url_table, help=f"Excel file with a url column. Default: {default_url_table}")
    parser.add_argument("--delay", type=float, default=2.0, help="Delay after opening each URL. Default: 2")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help=f"Output file. Supports .xlsx, .json, or .csv. Default: {DEFAULT_OUTPUT}")
    parser.add_argument("--headless", action="store_true", help="Run Chrome headlessly")
    args = parser.parse_args()

    input_path = resolve_path(args.url_table)
    output_path = resolve_path(args.output)
    urls = read_urls_from_excel(input_path)

    print("Mode: public pages from Excel, no login")
    print(f"Working directory: {Path.cwd()}")
    print(f"URL table: {input_path}")
    print(f"Output: {output_path}")
    print(f"URL count: {len(urls)}")

    if not urls:
        raise SystemExit("No URLs found in Excel file.")

    posts = scrape_urls(urls=urls, delay=args.delay, headless=args.headless)

    suffix = output_path.suffix.lower()
    if suffix == ".xlsx":
        save_excel(posts, output_path)
    elif suffix == ".csv":
        save_csv(posts, output_path)
    else:
        save_json(posts, output_path)

    ok_count = sum(1 for post in posts if not post.get("error"))
    print(f"Done: scraped {ok_count}/{len(posts)} URLs and saved to {output_path}")


if __name__ == "__main__":
    main()

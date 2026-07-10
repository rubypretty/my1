import argparse
import csv
import json
import os
import re
import time
from datetime import datetime, timedelta
from pathlib import Path
from urllib.parse import quote, urlsplit, urlunsplit
from urllib.request import urlopen

from common import build_driver
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from selenium import webdriver
from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from websocket import create_connection

from config import E_DATE, MAX_N, SEARCH_WORD, S_DATE


BASE_DIR = Path(__file__).resolve().parent
ENV_FILE = BASE_DIR / ".env"
DEFAULT_PROFILE_DIR = "chrome_profile"
DEFAULT_SEARCH_WORD = SEARCH_WORD
DEFAULT_S_DATE = S_DATE
DEFAULT_E_DATE = E_DATE
DEFAULT_MAX_N = MAX_N
OUTPUT_TIME_FORMAT = "%Y.%m.%d %H:%M"
VIEW_COUNT_XPATHS = [
    "/html/body/div[1]/div/div/div[3]/div[3]/div/div/div/div[1]/div[1]/div/div/div[1]/div[5]/div[1]/div/div/div[2]/span/span",
    "//span[contains(normalize-space(.), '\u6b21\u700f\u89bd')]",
    "//span[contains(translate(normalize-space(.), 'VIEWS', 'views'), 'views')]",
]
ACTION_COUNT_XPATHS = {
    "愛心": "/html/body/div[1]/div/div/div[3]/div[3]/div/div/div/div[1]/div[1]/div/div/div[2]/div/div[1]/div[1]/div/div[1]/div/div/div/div[3]/div/div[3]/div/div[1]/div/div/div/span/div/span",
    "留言": "/html/body/div[1]/div/div/div[3]/div[3]/div/div/div/div[1]/div[1]/div/div/div[2]/div/div[1]/div[1]/div/div[1]/div/div/div/div[3]/div/div[3]/div/div[2]/div/div/div/span/div/span",
    "轉發": "/html/body/div[1]/div/div/div[3]/div[3]/div/div/div/div[2]/div[1]/div/div/div[2]/div/div[1]/div[1]/div/div[1]/div/div/div/div[3]/div/div[3]/div/div[3]/div/div/div/span/div/span",
    "分享": "/html/body/div[1]/div/div/div[3]/div[3]/div/div/div/div[1]/div[1]/div/div/div[2]/div/div[1]/div[1]/div/div[1]/div/div/div/div[3]/div/div[2]/div/div[4]/div/div/div/span/div/span",
}


def load_local_env(path: Path = ENV_FILE) -> None:
    if not path.exists():
        return

    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue

        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def get_search_settings() -> tuple[str, str, str, int]:
    search_word = DEFAULT_SEARCH_WORD
    s_date = DEFAULT_S_DATE
    e_date = DEFAULT_E_DATE

    try:
        max_n = int(DEFAULT_MAX_N)
    except ValueError as error:
        raise SystemExit("MAX_N in config.py must be an integer.") from error

    return search_word, s_date, e_date, max_n


def normalize_post_url(url: str) -> str:
    parts = urlsplit(url.replace("https://www.threads.net/", "https://www.threads.com/"))
    path = parts.path.rstrip("/")
    return urlunsplit(("https", "www.threads.com", path, "hl=zh-tw", ""))


def is_media_post_url(url: str) -> bool:
    parts = urlsplit(url.replace("https://www.threads.net/", "https://www.threads.com/"))
    return parts.path.rstrip("/").endswith("/media")


def parse_date(date_text: str) -> datetime:
    for date_format in ("%Y%m%d", "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(date_text, date_format)
        except ValueError:
            pass
    raise ValueError(f"Invalid date: {date_text}. Use YYYYMMDD, YYYY-MM-DD, YYYY/MM/DD, or YYYY.MM.DD.")


def safe_filename_part(text: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", str(text).strip())
    cleaned = re.sub(r"\s+", "_", cleaned).strip(" ._")
    return cleaned or "keyword"


def default_posts_output_path(
    s_date: datetime,
    e_date: datetime,
    keyword: str,
    run_time: datetime | None = None,
) -> Path:
    run_time = run_time or datetime.now().astimezone()
    filename = (
        f"{s_date:%Y%m%d}_{e_date:%Y%m%d}_"
        f"{safe_filename_part(keyword)}_{run_time:%Y%m%d_%H%M}.xlsx"
    )
    return Path(filename)


def date_markers(day: datetime) -> set[str]:
    return {
        day.strftime("%Y-%m-%d"),
        day.strftime("%Y%m%d"),
        day.strftime("%Y/%m/%d"),
        day.strftime("%Y.%m.%d"),
        f"{day.year}-{day.month}-{day.day}",
        f"{day.year}/{day.month}/{day.day}",
        f"{day.year}.{day.month}.{day.day}",
    }


def date_markers_in_range(start_date: datetime, end_date: datetime) -> set[str]:
    markers = set()
    current = start_date
    while current <= end_date:
        markers.update(date_markers(current))
        current += timedelta(days=1)
    return markers


def looks_in_date_range(text: str, start_date: datetime, end_date: datetime) -> bool:
    if any(marker in text for marker in date_markers_in_range(start_date, end_date)):
        return True

    today = datetime.now().astimezone().replace(tzinfo=None)
    includes_today = start_date.date() <= today.date() <= end_date.date()
    if not includes_today:
        return False

    relative_patterns = [
        r"\b\d+\s*s\b",
        r"\b\d+\s*m\b",
        r"\b\d+\s*h\b",
        r"\b\d+\s*sec\b",
        r"\b\d+\s*min\b",
        r"\b\d+\s*hr\b",
        r"\d+\s*秒",
        r"\d+\s*分鐘",
        r"\d+\s*分",
        r"\d+\s*小時",
        r"剛剛",
        r"現在",
    ]
    return any(re.search(pattern, text, flags=re.IGNORECASE) for pattern in relative_patterns)


def wait_for_body(driver: webdriver.Chrome, timeout: int = 20) -> str:
    wait = WebDriverWait(driver, timeout)
    wait.until(lambda current: current.find_elements(By.TAG_NAME, "body"))
    wait.until(lambda current: len((current.find_element(By.TAG_NAME, "body").text or "").strip()) > 20)
    return driver.find_element(By.TAG_NAME, "body").text or ""


def normalize_view_count(text: str) -> str:
    return " ".join(text.replace("\xa0", " ").split())


def format_count_number(text: str) -> str:
    normalized = normalize_view_count(text)
    match = re.search(r"(\d[\d,]*(?:\.\d+)?)\s*(億|萬|千|k|m|b)?", normalized, flags=re.IGNORECASE)
    if not match:
        return normalized

    amount = float(match.group(1).replace(",", ""))
    unit = (match.group(2) or "").lower()
    multipliers = {
        "千": 1_000,
        "萬": 10_000,
        "億": 100_000_000,
        "k": 1_000,
        "m": 1_000_000,
        "b": 1_000_000_000,
    }
    value = int(round(amount * multipliers.get(unit, 1)))
    return f"{value:,}"


def format_view_count(text: str) -> str:
    normalized = normalize_view_count(text)
    return format_count_number(normalized) if looks_like_view_count(normalized) else normalized


def looks_like_view_count(text: str) -> bool:
    normalized = normalize_view_count(text).lower()
    return bool(normalized) and (
        "\u6b21\u700f\u89bd" in normalized
        or "\u700f\u89bd\u6b21\u6578" in normalized
        or "view" in normalized
    )


def extract_view_count_from_dom(driver: webdriver.Chrome) -> str:
    for xpath in VIEW_COUNT_XPATHS:
        for element in driver.find_elements(By.XPATH, xpath):
            text = normalize_view_count(element.text or element.get_attribute("textContent") or "")
            if looks_like_view_count(text):
                return format_view_count(text)

    body_text = driver.find_element(By.TAG_NAME, "body").text or ""
    for line in body_text.splitlines():
        if looks_like_view_count(line):
            return format_view_count(line)
    return ""


def empty_action_counts() -> dict[str, str]:
    return {"愛心": "", "留言": "", "轉發": "", "分享": ""}


def normalize_action_count(text: str) -> str:
    return " ".join(text.replace("\xa0", " ").split())


def extract_action_counts_from_dom(driver: webdriver.Chrome) -> dict[str, str]:
    xpath_counts = empty_action_counts()
    for action, xpath in ACTION_COUNT_XPATHS.items():
        for element in driver.find_elements(By.XPATH, xpath):
            text = normalize_action_count(element.text or element.get_attribute("textContent") or "")
            if text:
                xpath_counts[action] = text
                break

    script = r"""
    const normalize = (text) => (text || '').replace(/\u00a0/g, ' ').replace(/\s+/g, ' ').trim();
    const numberFrom = (text) => {
        const match = normalize(text).match(/(\d[\d,]*(?:\.\d+)?\s*(?:萬|千|億|k|m|b)?)/i);
        return match ? match[1].replace(/\s+/g, '') : '';
    };
    const actionFor = (text) => {
        const value = normalize(text).toLowerCase();
        if (!value) return '';
        if (/(un)?like|likes?|讚|喜歡|愛心/.test(value)) return '愛心';
        if (/repl(y|ies)|comments?|回覆|留言/.test(value)) return '留言';
        if (/repost|reposts|reshare|quote|轉發|轉貼|引用/.test(value)) return '轉發';
        if (/share|shares|分享/.test(value)) return '分享';
        return '';
    };
    const labelsFor = (element) => [
        element.getAttribute('aria-label'),
        element.getAttribute('title'),
        element.getAttribute('data-testid'),
        element.innerText,
        element.textContent,
    ].map(normalize).filter(Boolean);
    const result = {'愛心': '', '留言': '', '轉發': '', '分享': ''};
    const selectors = [
        '[role="button"]',
        'button',
        'a[aria-label]',
        'div[aria-label]',
        'span[aria-label]',
    ].join(',');
    for (const element of document.querySelectorAll(selectors)) {
        const labels = labelsFor(element);
        const action = labels.map(actionFor).find(Boolean);
        if (!action || result[action]) continue;

        const directNumber = labels.map(numberFrom).find(Boolean);
        if (directNumber) {
            result[action] = directNumber;
            continue;
        }

        const siblingTexts = [];
        if (element.nextElementSibling) siblingTexts.push(element.nextElementSibling.innerText || element.nextElementSibling.textContent);
        if (element.previousElementSibling) siblingTexts.push(element.previousElementSibling.innerText || element.previousElementSibling.textContent);
        for (const text of siblingTexts) {
            const number = numberFrom(text);
            if (number) {
                result[action] = number;
                break;
            }
        }
    }
    return result;
    """
    counts = driver.execute_script(script) or {}
    fallback_counts = {key: str(value) for key, value in counts.items() if value}
    return {
        action: xpath_counts[action] or fallback_counts.get(action, "")
        for action in empty_action_counts()
    }


def clean_lines(text: str) -> list[str]:
    stop_markers = [
        "登入即可查看更多",
        "登入查看更多 Threads",
        "© 2026",
        "Threads 使用條款",
    ]
    for marker in stop_markers:
        if marker in text:
            text = text.split(marker, 1)[0].strip()

    noise_lines = {"登入", "註冊", "搜尋", "Threads"}
    lines = []
    for line in text.splitlines():
        line = line.strip()
        if not line or line in noise_lines:
            continue
        lines.append(line)
    return lines


def extract_author_from_url(url: str) -> str:
    match = re.search(r"/@([^/?#]+)/post/", url)
    return match.group(1) if match else ""


def looks_like_metric(line: str) -> bool:
    return bool(re.fullmatch(r"\d[\d,]*", line))


def looks_like_time(line: str) -> bool:
    if re.fullmatch(r"\d+\s*(秒|分鐘|小時|天|週|月|年|s|m|h|sec|min|hr|d|w)", line, flags=re.IGNORECASE):
        return True
    if re.fullmatch(r"\d{4}-\d{1,2}-\d{1,2}", line):
        return True
    return line in {"昨天", "前天"}


def format_post_time(post_time: str, reference_time: datetime | None = None) -> str:
    if not post_time:
        return ""

    reference_time = reference_time or datetime.now().astimezone()
    absolute_match = re.fullmatch(r"(\d{4})-(\d{1,2})-(\d{1,2})", post_time)
    if absolute_match:
        year, month, day = (int(part) for part in absolute_match.groups())
        return datetime(year, month, day, tzinfo=reference_time.tzinfo).strftime(OUTPUT_TIME_FORMAT)

    relative_match = re.fullmatch(
        r"(\d+)\s*(秒|分鐘|小時|天|週|月|年|s|m|h|sec|min|hr|d|w)",
        post_time,
        flags=re.IGNORECASE,
    )
    if relative_match:
        amount = int(relative_match.group(1))
        unit = relative_match.group(2).lower()
        deltas = {
            "秒": timedelta(seconds=amount),
            "s": timedelta(seconds=amount),
            "sec": timedelta(seconds=amount),
            "分鐘": timedelta(minutes=amount),
            "m": timedelta(minutes=amount),
            "min": timedelta(minutes=amount),
            "小時": timedelta(hours=amount),
            "h": timedelta(hours=amount),
            "hr": timedelta(hours=amount),
            "天": timedelta(days=amount),
            "d": timedelta(days=amount),
            "週": timedelta(weeks=amount),
            "w": timedelta(weeks=amount),
            "月": timedelta(days=amount * 30),
            "年": timedelta(days=amount * 365),
        }
        return (reference_time - deltas[unit]).strftime(OUTPUT_TIME_FORMAT)

    if post_time == "昨天":
        return (reference_time - timedelta(days=1)).strftime(OUTPUT_TIME_FORMAT)
    if post_time == "前天":
        return (reference_time - timedelta(days=2)).strftime(OUTPUT_TIME_FORMAT)
    return post_time


def drop_carousel_markers(lines: list[str]) -> list[str]:
    cleaned = []
    index = 0
    while index < len(lines):
        if (
            index + 2 < len(lines)
            and lines[index].isdigit()
            and lines[index + 1] == "/"
            and lines[index + 2].isdigit()
        ):
            index += 3
            continue
        cleaned.append(lines[index])
        index += 1
    return cleaned


def find_reply_start(lines: list[str], start: int) -> int | None:
    for index in range(start, len(lines)):
        line = lines[index]
        if line.startswith("回覆"):
            return index + 1
        if line in {"回覆", "更多回覆"}:
            return index
    return None


def find_metric_start(lines: list[str], start: int, stop: int) -> int | None:
    for index in range(start, stop):
        if looks_like_metric(lines[index]):
            metric_count = 0
            for line in lines[index : min(stop, index + 6)]:
                if looks_like_metric(line):
                    metric_count += 1
            if metric_count >= 2:
                return index
    return None


def find_metric_end(lines: list[str], start: int, stop: int) -> int:
    index = start
    while index < stop and looks_like_metric(lines[index]):
        index += 1
    return index


def trim_ui_lines(lines: list[str]) -> list[str]:
    return [
        line
        for line in lines
        if line not in {"回覆", "更多回覆", "查看翻譯"} and not line.startswith("回覆")
    ]


def parse_thread_lines(
    lines: list[str],
    source_url: str = "",
    reference_time: datetime | None = None,
) -> dict:
    if not lines:
        return {}

    lines = drop_carousel_markers(lines)
    author_from_url = extract_author_from_url(source_url)

    view_index = next((index for index, line in enumerate(lines) if looks_like_view_count(line)), None)
    view_count = format_view_count(lines[view_index]) if view_index is not None else ""

    author_index = None
    if author_from_url:
        author_index = next((index for index, line in enumerate(lines) if line == author_from_url), None)
    if author_index is None:
        start = (view_index + 1) if view_index is not None else 0
        author_index = start if start < len(lines) else None

    author = author_from_url or (lines[author_index] if author_index is not None else "")

    post_time = ""
    time_index = None
    if author_index is not None:
        for index in range(author_index + 1, min(len(lines), author_index + 6)):
            if looks_like_time(lines[index]):
                post_time = lines[index]
                time_index = index
                break

    text_start = (time_index + 1) if time_index is not None else ((author_index + 1) if author_index is not None else 0)
    reply_start = find_reply_start(lines, text_start)
    main_stop = reply_start - 1 if reply_start is not None else len(lines)
    metric_start = find_metric_start(lines, text_start, main_stop)
    if metric_start is not None:
        main_stop = metric_start

    main_lines = trim_ui_lines(lines[text_start:main_stop])

    metrics_stop = reply_start - 1 if reply_start is not None else len(lines)
    metric_source_start = metric_start if metric_start is not None else text_start
    metric_end = find_metric_end(lines, metric_source_start, metrics_stop)
    metrics = [
        line
        for line in lines[metric_source_start:metric_end]
        if looks_like_metric(line)
    ]

    if reply_start is not None:
        extra_lines = trim_ui_lines(lines[reply_start:])
    elif metric_start is not None:
        extra_lines = trim_ui_lines(lines[metric_end:])
    else:
        extra_lines = []

    return {
        "author": author,
        "post_time": format_post_time(post_time, reference_time),
        "view_count": view_count,
        "main_text": "\n".join(main_lines).strip(),
        "愛心": metrics[0] if len(metrics) > 0 else "",
        "留言": metrics[1] if len(metrics) > 1 else "",
        "轉發": metrics[2] if len(metrics) > 2 else "",
        "分享": "",
        "related_or_replies": "\n".join(extra_lines).strip(),
    }


def search_url(keyword: str) -> str:
    return f"https://www.threads.com/search?q={quote(keyword)}"


def collect_post_urls(driver: webdriver.Chrome, keyword: str, max_scrolls: int, delay: float) -> list[str]:
    driver.get(search_url(keyword))
    time.sleep(delay)

    urls: list[str] = []
    seen = set()
    for _ in range(max_scrolls):
        try:
            wait_for_body(driver)
        except TimeoutException:
            pass

        hrefs = driver.execute_script(
            "return Array.from(document.querySelectorAll(\"a[href*='/post/']\")).map((a) => a.href);"
        ) or []
        for href in hrefs:
            if not href or "threads." not in href:
                continue
            if is_media_post_url(href):
                continue
            normalized = normalize_post_url(href)
            if normalized in seen:
                continue
            seen.add(normalized)
            urls.append(normalized)

        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(delay)

    return urls


def extract_post(
    driver: webdriver.Chrome,
    url: str,
    keyword: str,
    start_date: datetime,
    end_date: datetime,
    delay: float,
) -> dict | None:
    if is_media_post_url(url):
        return None

    driver.get(url.replace("https://www.threads.com/", "https://www.threads.net/"))
    time.sleep(delay)

    try:
        body_text = wait_for_body(driver)
    except TimeoutException:
        return None

    if keyword not in body_text:
        return None
    if not looks_in_date_range(body_text, start_date, end_date):
        return None

    return {
        "search_word": keyword,
        "url": normalize_post_url(url),
        "final_url": driver.current_url,
        "text": body_text,
        "raw_html": driver.page_source or "",
        "view_count": extract_view_count_from_dom(driver),
        "action_counts": extract_action_counts_from_dom(driver),
        "matched_date_range": f"{start_date:%Y-%m-%d}~{end_date:%Y-%m-%d}",
        "scraped_at": datetime.now().astimezone().isoformat(timespec="seconds"),
    }


def save_posts_json(posts: list[dict], output: Path) -> None:
    output.write_text(json.dumps(posts, ensure_ascii=False, indent=2), encoding="utf-8")


def save_posts_csv(posts: list[dict], output: Path) -> None:
    fieldnames = [
        "num",
        "source_url",
        "final_url",
        "author",
        "post_time",
        "view_count",
        "main_text",
        "愛心",
        "留言",
        "轉發",
        "分享",
        "related_or_replies",
        "error",
        "scraped_at",
    ]
    with output.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(posts)


def save_posts_excel(posts: list[dict], output: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "posts"
    raw = workbook.create_sheet("raw")

    summary_headers = [
        "num",
        "source_url",
        "author",
        "post_time",
        "view_count",
        "main_text",
        "愛心",
        "留言",
        "轉發",
        "分享",
        "related_or_replies",
        "error",
        "scraped_at",
    ]
    raw_headers = ["source_url", "final_url", "text", "error", "scraped_at"]

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for sheet, headers in [(summary, summary_headers), (raw, raw_headers)]:
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for post in posts:
        summary.append([post.get(header, "") for header in summary_headers])
        summary_url_cell = summary.cell(row=summary.max_row, column=summary_headers.index("source_url") + 1)
        if summary_url_cell.value:
            summary_url_cell.hyperlink = summary_url_cell.value
            summary_url_cell.style = "Hyperlink"

        raw.append([post.get(header, "") for header in raw_headers])
        raw_url_cell = raw.cell(row=raw.max_row, column=raw_headers.index("source_url") + 1)
        if raw_url_cell.value:
            raw_url_cell.hyperlink = raw_url_cell.value
            raw_url_cell.style = "Hyperlink"

    for sheet in [summary, raw]:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column, width in {
        "A": 10,
        "B": 56,
        "C": 20,
        "D": 14,
        "E": 14,
        "F": 60,
        "G": 16,
        "H": 16,
        "I": 16,
        "J": 16,
        "K": 70,
        "L": 32,
        "M": 22,
    }.items():
        summary.column_dimensions[column].width = width

    raw.column_dimensions["A"].width = 56
    raw.column_dimensions["B"].width = 56
    raw.column_dimensions["C"].width = 100
    raw.column_dimensions["D"].width = 32
    raw.column_dimensions["E"].width = 22
    workbook.save(output)


def build_downloaded_post(post: dict) -> dict:
    scraped_at = datetime.now().astimezone()
    text = (post.get("text") or "").strip()
    lines = clean_lines(text)
    cleaned_text = "\n".join(lines).strip()
    parsed = parse_thread_lines(
        lines,
        source_url=post.get("url", ""),
        reference_time=scraped_at,
    )
    if post.get("view_count"):
        parsed["view_count"] = post["view_count"]
    for key, value in (post.get("action_counts") or {}).items():
        if value:
            parsed[key] = value

    return {
        "source_url": post.get("url", ""),
        "final_url": post.get("final_url", ""),
        **parsed,
        "text": cleaned_text or text,
        "raw_html": post.get("raw_html", post.get("raw_xml", "")),
        "error": "" if cleaned_text else "No readable public post text found.",
        "scraped_at": scraped_at.strftime("%Y-%m-%dT%H:%M:%S%z"),
    }


def build_error_post(url: str, error: Exception | str) -> dict:
    scraped_at = datetime.now().astimezone()
    return {
        "source_url": normalize_post_url(url) if url else "",
        "final_url": "",
        "author": extract_author_from_url(url),
        "post_time": "",
        "view_count": "",
        "main_text": "",
        "愛心": "",
        "留言": "",
        "轉發": "",
        "分享": "",
        "related_or_replies": "",
        "text": "",
        "raw_html": "",
        "error": f"{type(error).__name__}: {error}" if isinstance(error, Exception) else str(error),
        "scraped_at": scraped_at.strftime("%Y-%m-%dT%H:%M:%S%z"),
    }


def make_browser_openable_html(raw_html: str) -> str:
    html = raw_html or "<!doctype html><html><body></body></html>"
    stripped = html.lstrip().lower()
    if not stripped.startswith("<!doctype") and stripped.startswith("<html"):
        html = "<!doctype html>\n" + html
    return html


def save_raw_html_files(posts: list[dict], output: Path) -> None:
    html_dir = output.with_suffix("")
    html_dir.mkdir(parents=True, exist_ok=True)
    for post in posts:
        raw_html = post.get("raw_html", post.get("raw_xml", ""))
        filename = f"{int(post.get('num', 0)):04d}.html"
        (html_dir / filename).write_text(make_browser_openable_html(raw_html), encoding="utf-8")


def save_downloaded_posts(posts: list[dict], output: Path) -> None:
    numbered_posts = [
        {**post, "num": index}
        for index, post in enumerate(posts, start=1)
    ]
    output_posts = [
        {key: value for key, value in post.items() if key not in {"raw_html", "raw_xml"}}
        for post in numbered_posts
    ]
    suffix = output.suffix.lower()
    if suffix == ".xlsx":
        save_posts_excel(output_posts, output)
    elif suffix == ".csv":
        save_posts_csv(output_posts, output)
    else:
        save_posts_json(output_posts, output)
    save_raw_html_files(numbered_posts, output)


def read_debugger_address(profile_dir: str | Path) -> str | None:
    devtools_file = Path(profile_dir) / "DevToolsActivePort"
    if not devtools_file.exists():
        return None

    lines = devtools_file.read_text(encoding="utf-8").splitlines()
    if not lines:
        return None
    port = lines[0].strip()
    return f"127.0.0.1:{port}" if port else None


class CdpPage:
    def __init__(self, debugger_address: str):
        self.debugger_address = debugger_address
        self.ws = create_connection(self.get_page_websocket_url())
        self.next_id = 0
        self.call("Page.enable")
        self.call("Runtime.enable")

    def get_page_websocket_url(self) -> str:
        with urlopen(f"http://{self.debugger_address}/json/list", timeout=5) as response:
            targets = json.loads(response.read().decode("utf-8"))

        pages = [target for target in targets if target.get("type") == "page"]
        threads_pages = [target for target in pages if "threads." in target.get("url", "")]
        target = threads_pages[0] if threads_pages else pages[0]
        return target["webSocketDebuggerUrl"]

    def call(self, method: str, params: dict | None = None) -> dict:
        self.next_id += 1
        message = {"id": self.next_id, "method": method}
        if params is not None:
            message["params"] = params
        self.ws.send(json.dumps(message))

        while True:
            response = json.loads(self.ws.recv())
            if response.get("id") == self.next_id:
                if "error" in response:
                    raise RuntimeError(response["error"])
                return response.get("result", {})

    def eval(self, expression: str):
        result = self.call(
            "Runtime.evaluate",
            {
                "expression": expression,
                "returnByValue": True,
                "awaitPromise": True,
            },
        )
        return result.get("result", {}).get("value")

    def navigate(self, url: str, delay: float) -> None:
        self.call("Page.navigate", {"url": url})
        time.sleep(delay)

    def wait_for_body(self, timeout: int = 20) -> str:
        deadline = time.time() + timeout
        while time.time() < deadline:
            text = self.eval("document.body ? document.body.innerText : ''") or ""
            if len(text.strip()) > 20:
                return text
            time.sleep(0.5)
        raise TimeoutException("Timed out waiting for body text.")

    def close(self) -> None:
        self.ws.close()


def collect_post_urls_cdp(page: CdpPage, keyword: str, max_scrolls: int, delay: float) -> list[str]:
    page.navigate(search_url(keyword), delay=delay)

    urls: list[str] = []
    seen = set()
    for _ in range(max_scrolls):
        try:
            page.wait_for_body()
        except TimeoutException:
            pass

        hrefs = page.eval(
            "Array.from(document.querySelectorAll(\"a[href*='/post/']\")).map((a) => a.href)"
        ) or []
        for href in hrefs:
            if not href or "threads." not in href:
                continue
            if is_media_post_url(href):
                continue
            normalized = normalize_post_url(href)
            if normalized in seen:
                continue
            seen.add(normalized)
            urls.append(normalized)

        page.eval("window.scrollTo(0, document.body.scrollHeight)")
        time.sleep(delay)

    return urls


def extract_view_count_cdp(page: CdpPage) -> str:
    text = page.eval(
        """
        (() => {
            const normalize = (text) => (text || '').replace(/\\u00a0/g, ' ').replace(/\\s+/g, ' ').trim();
            const xpaths = [
                "/html/body/div[1]/div/div/div[3]/div[3]/div/div/div/div[1]/div[1]/div/div/div[1]/div[5]/div[1]/div/div/div[2]/span/span",
                "//span[contains(normalize-space(.), '\\u6b21\\u700f\\u89bd')]",
                "//span[contains(translate(normalize-space(.), 'VIEWS', 'views'), 'views')]",
            ];
            for (const xpath of xpaths) {
                const result = document.evaluate(xpath, document, null, XPathResult.ORDERED_NODE_SNAPSHOT_TYPE, null);
                for (let index = 0; index < result.snapshotLength; index += 1) {
                    const text = normalize(result.snapshotItem(index).textContent);
                    const lower = text.toLowerCase();
                    if (text && (text.includes('\\u6b21\\u700f\\u89bd') || text.includes('\\u700f\\u89bd\\u6b21\\u6578') || lower.includes('view'))) {
                        return text;
                    }
                }
            }
            const bodyLines = (document.body ? document.body.innerText : '').split('\\n').map(normalize);
            for (const line of bodyLines) {
                const lower = line.toLowerCase();
                if (line && (line.includes('\\u6b21\\u700f\\u89bd') || line.includes('\\u700f\\u89bd\\u6b21\\u6578') || lower.includes('view'))) {
                    return line;
                }
            }
            return '';
        })()
        """
    ) or ""
    return format_view_count(text) if text else ""


def extract_action_counts_cdp(page: CdpPage) -> dict[str, str]:
    counts = page.eval(
        r"""
        (() => {
            const normalize = (text) => (text || '').replace(/\u00a0/g, ' ').replace(/\s+/g, ' ').trim();
            const xpathMap = {
                '愛心': "/html/body/div[1]/div/div/div[3]/div[3]/div/div/div/div[1]/div[1]/div/div/div[2]/div/div[1]/div[1]/div/div[1]/div/div/div/div[3]/div/div[3]/div/div[1]/div/div/div/span/div/span",
                '留言': "/html/body/div[1]/div/div/div[3]/div[3]/div/div/div/div[1]/div[1]/div/div/div[2]/div/div[1]/div[1]/div/div[1]/div/div/div/div[3]/div/div[3]/div/div[2]/div/div/div/span/div/span",
                '轉發': "/html/body/div[1]/div/div/div[3]/div[3]/div/div/div/div[2]/div[1]/div/div/div[2]/div/div[1]/div[1]/div/div[1]/div/div/div/div[3]/div/div[3]/div/div[3]/div/div/div/span/div/span",
                '分享': "/html/body/div[1]/div/div/div[3]/div[3]/div/div/div/div[1]/div[1]/div/div/div[2]/div/div[1]/div[1]/div/div[1]/div/div/div/div[3]/div/div[2]/div/div[4]/div/div/div/span/div/span",
            };
            const numberFrom = (text) => {
                const match = normalize(text).match(/(\d[\d,]*(?:\.\d+)?\s*(?:萬|千|億|k|m|b)?)/i);
                return match ? match[1].replace(/\s+/g, '') : '';
            };
            const actionFor = (text) => {
                const value = normalize(text).toLowerCase();
                if (!value) return '';
                if (/(un)?like|likes?|讚|喜歡|愛心/.test(value)) return '愛心';
                if (/repl(y|ies)|comments?|回覆|留言/.test(value)) return '留言';
                if (/repost|reposts|reshare|quote|轉發|轉貼|引用/.test(value)) return '轉發';
                if (/share|shares|分享/.test(value)) return '分享';
                return '';
            };
            const labelsFor = (element) => [
                element.getAttribute('aria-label'),
                element.getAttribute('title'),
                element.getAttribute('data-testid'),
                element.innerText,
                element.textContent,
            ].map(normalize).filter(Boolean);
            const result = {'愛心': '', '留言': '', '轉發': '', '分享': ''};
            const selectors = [
                '[role="button"]',
                'button',
                'a[aria-label]',
                'div[aria-label]',
                'span[aria-label]',
            ].join(',');
            for (const [action, xpath] of Object.entries(xpathMap)) {
                const xpathResult = document.evaluate(xpath, document, null, XPathResult.ORDERED_NODE_SNAPSHOT_TYPE, null);
                for (let index = 0; index < xpathResult.snapshotLength; index += 1) {
                    const text = normalize(xpathResult.snapshotItem(index).textContent);
                    if (text) {
                        result[action] = text;
                        break;
                    }
                }
            }
            for (const element of document.querySelectorAll(selectors)) {
                const labels = labelsFor(element);
                const action = labels.map(actionFor).find(Boolean);
                if (!action || result[action]) continue;

                const directNumber = labels.map(numberFrom).find(Boolean);
                if (directNumber) {
                    result[action] = directNumber;
                    continue;
                }

                const siblingTexts = [];
                if (element.nextElementSibling) siblingTexts.push(element.nextElementSibling.innerText || element.nextElementSibling.textContent);
                if (element.previousElementSibling) siblingTexts.push(element.previousElementSibling.innerText || element.previousElementSibling.textContent);
                for (const text of siblingTexts) {
                    const number = numberFrom(text);
                    if (number) {
                        result[action] = number;
                        break;
                    }
                }
            }
            return result;
        })()
        """
    ) or {}
    return {**empty_action_counts(), **{key: str(value) for key, value in counts.items() if value}}


def extract_post_cdp(
    page: CdpPage,
    url: str,
    keyword: str,
    start_date: datetime,
    end_date: datetime,
    delay: float,
) -> dict | None:
    if is_media_post_url(url):
        return None

    page.navigate(url.replace("https://www.threads.com/", "https://www.threads.net/"), delay=delay)

    try:
        body_text = page.wait_for_body()
    except TimeoutException:
        return None

    if keyword not in body_text:
        return None
    if not looks_in_date_range(body_text, start_date, end_date):
        return None

    final_url = page.eval("location.href") or url
    return {
        "search_word": keyword,
        "url": normalize_post_url(url),
        "final_url": final_url,
        "text": body_text,
        "raw_html": page.eval("document.documentElement ? document.documentElement.outerHTML : ''") or "",
        "view_count": extract_view_count_cdp(page),
        "action_counts": extract_action_counts_cdp(page),
        "matched_date_range": f"{start_date:%Y-%m-%d}~{end_date:%Y-%m-%d}",
        "scraped_at": datetime.now().astimezone().isoformat(timespec="seconds"),
    }


def run_with_cdp(
    debugger_address: str,
    keyword: str,
    start_date: datetime,
    end_date: datetime,
    limit: int,
    max_scrolls: int,
    delay: float,
    posts_output: Path,
) -> bool:
    try:
        page = CdpPage(debugger_address)
    except Exception as error:
        print(f"Could not use existing Chrome through CDP: {error}")
        return False

    try:
        print(f"Using existing Chrome through CDP: {debugger_address}")
        candidate_urls = collect_post_urls_cdp(page, keyword, max_scrolls=max_scrolls, delay=delay)
        print(f"Candidate posts: {len(candidate_urls)}")

        posts = []
        downloaded_posts = []
        for url in candidate_urls:
            if len(posts) >= limit:
                break
            print(f"Checking: {url}")
            try:
                post = extract_post_cdp(
                    page,
                    url,
                    keyword=keyword,
                    start_date=start_date,
                    end_date=end_date,
                    delay=delay,
                )
                if post:
                    posts.append(post)
                    downloaded_post = build_downloaded_post(post)
                    downloaded_posts.append(downloaded_post)
                    save_downloaded_posts(downloaded_posts, posts_output)
                    print(f"  kept: {len(posts)}/{limit}")
                    print(f"  downloaded: {posts_output.resolve()}")
            except Exception as error:
                downloaded_posts.append(build_error_post(url, error))
                save_downloaded_posts(downloaded_posts, posts_output)
                print(f"  error recorded: {error}")

        save_downloaded_posts(downloaded_posts, posts_output)
        print(f"Saved downloaded posts: {posts_output.resolve()}")
        return True
    finally:
        page.close()


def start_driver(profile_dir: str | Path) -> webdriver.Chrome:
    debugger_address = read_debugger_address(profile_dir)
    if debugger_address:
        try:
            print(f"Attaching to existing Chrome: {debugger_address}")
            return build_driver(headless=False, debugger_address=debugger_address)
        except WebDriverException:
            print("Could not attach to existing Chrome. Starting a new Chrome session.")

    return build_driver(headless=False, user_data_dir=profile_dir)


def main() -> None:
    load_local_env()
    search_word, s_date, e_date, max_n = get_search_settings()

    parser = argparse.ArgumentParser(description="Search Threads posts by keyword and date range.")
    parser.add_argument("--keyword", default=search_word, help=f"Search keyword. Default: {search_word}")
    parser.add_argument("--s-date", default=s_date, help=f"Search start date. Default: {s_date}")
    parser.add_argument("--e-date", default=e_date, help=f"Search end date. Default: {e_date}")
    parser.add_argument("--limit", "--max-n", dest="limit", type=int, default=max_n, help=f"Maximum posts to keep. Default: {max_n}")
    parser.add_argument("--profile-dir", default=DEFAULT_PROFILE_DIR, help=f"Chrome profile directory. Default: {DEFAULT_PROFILE_DIR}")
    parser.add_argument("--posts-output", default=None, help="Downloaded posts output path. Default: <S_DATE>_<E_DATE>_<SEARCH_WORD>_<run_time>.xlsx")
    parser.add_argument("--delay", type=float, default=2.0, help="Delay after page actions. Default: 2")
    parser.add_argument("--max-scrolls", type=int, default=12, help="Maximum search result scrolls. Default: 12")
    args = parser.parse_args()

    start_date = parse_date(args.s_date)
    end_date = parse_date(args.e_date)
    if start_date > end_date:
        raise SystemExit("s_date must be earlier than or equal to e_date.")

    posts_output = Path(args.posts_output) if args.posts_output else default_posts_output_path(
        start_date,
        end_date,
        args.keyword,
    )
    debugger_address = read_debugger_address(args.profile_dir)

    print(f"Search word: {args.keyword}")
    print(f"Date range: {start_date:%Y-%m-%d} ~ {end_date:%Y-%m-%d}")
    print(f"Max posts: {args.limit}")
    if debugger_address and run_with_cdp(
        debugger_address,
        keyword=args.keyword,
        start_date=start_date,
        end_date=end_date,
        limit=args.limit,
        max_scrolls=args.max_scrolls,
        delay=args.delay,
        posts_output=posts_output,
    ):
        return

    try:
        driver = start_driver(args.profile_dir)
    except WebDriverException as error:
        raise SystemExit(
            "Could not start Chrome with chrome_profile. Close any Chrome window using this profile and try again.\n"
            f"{error}"
        ) from error

    try:
        try:
            candidate_urls = collect_post_urls(driver, args.keyword, max_scrolls=args.max_scrolls, delay=args.delay)
        except Exception as error:
            save_downloaded_posts([build_error_post("", f"collect_post_urls failed: {type(error).__name__}: {error}")], posts_output)
            print(f"Candidate collection error recorded: {error}")
            return
        print(f"Candidate posts: {len(candidate_urls)}")

        posts = []
        downloaded_posts = []
        for url in candidate_urls:
            if len(posts) >= args.limit:
                break
            print(f"Checking: {url}")
            try:
                post = extract_post(
                    driver,
                    url,
                    keyword=args.keyword,
                    start_date=start_date,
                    end_date=end_date,
                    delay=args.delay,
                )
                if post:
                    posts.append(post)
                    downloaded_post = build_downloaded_post(post)
                    downloaded_posts.append(downloaded_post)
                    save_downloaded_posts(downloaded_posts, posts_output)
                    print(f"  kept: {len(posts)}/{args.limit}")
                    print(f"  downloaded: {posts_output.resolve()}")
            except Exception as error:
                downloaded_posts.append(build_error_post(url, error))
                save_downloaded_posts(downloaded_posts, posts_output)
                print(f"  error recorded: {error}")

        save_downloaded_posts(downloaded_posts, posts_output)
        print(f"Saved downloaded posts: {posts_output.resolve()}")
    finally:
        driver.quit()


if __name__ == "__main__":
    main()
